from tkinter import * 
from tkinter import ttk
import serial.tools.list_ports
import serial
import threading
import tkinter.filedialog
import re
import tkinter.messagebox as tm
import time
import os
from openpyxl import Workbook

excel_line = 2
fail_num = 0
filenames = ''
vfilenames = ''
ser = 0
ser1 = 0
comport_new = ''
bpsport_new = '115200'
comport_new1 = ''
bpsport_new1 = '115200'
top = 0
buadrate_list = ['9600','19200','38400','115200','230400']
log_filepath = 'F:\\IMU383_SPI_AUTOTEST'
t = []
test_steps = 1
result_com = ''

def packet_1M_check(file):
    global result_com,fail_num
    f = open(log_filepath+'\\spi_data_packet.txt','r')
    f_lines = f.readlines()
    f.close()
    for line in f_lines:
        if re.match('0000.*',line):
            line_list = line[:-1].split()
            burst_list_int = []
            for cell in line_list:
                burst_list_int.append(abs(float(cell)))
            if(0<=burst_list_int[1]<1 and 0<=burst_list_int[2]<1 and 0<=burst_list_int[3]<1 and 0<=burst_list_int[4]<0.1 and 0<=burst_list_int[5]<0.1 and 0.9<burst_list_int[6]<=1.1):
                result_com = '------PASS'
            else:
                result_com = '------FAIL'
                fail_num = fail_num + 1
                break

def default_check(text_handel,address,descript,f_log,fv_lines,index,type,ws):
    global ser,fail_num,test_steps,excel_line
    ser.write(bytes.fromhex(address))
    current_time = time.strftime('%Y_%m_%d_%H:%M:%S',time.localtime(time.time()))
    text_handel.insert(END, current_time+'    ' +'Step%d:'%test_steps+descript+'\n')
    f_log.write(current_time+'    '+'Step%d:'%test_steps+descript+'\n')
    test_steps += 1
    time.sleep(0.1)
    read = ser.readall()
    if len(read) > 0 or type == 'power_cycle' or type == 'write_only':
        current_time = time.strftime('%Y_%m_%d_%H:%M:%S',time.localtime(time.time()))
        if type == 'default_check1':
            if fv_lines[index][:-1] == bytes(read).decode('gb2312'):
                result_com = '------PASS'
            else:
                result_com = '------FAIL'
            ws.cell(row = excel_line,column = 1,value = descript)
            ws.cell(row = excel_line,column = 3,value = fv_lines[index][:-1])
            ws.cell(row = excel_line,column = 4,value = bytes(read).decode('gb2312'))
            ws.cell(row = excel_line,column = 5,value = result_com)
        elif type == 'read_only':
            if re.match(fv_lines[index][:-1]+'.{2}',bytes(read).decode('gb2312')):
                result_com = '------FAIL'
            else:
                result_com = '------PASS'
            ws.cell(row = excel_line,column = 1,value = descript)
            ws.cell(row = excel_line,column = 2,value = fv_lines[index][:-1])
            ws.cell(row = excel_line,column = 3,value = 'NOT'+fv_lines[index][:-1])
            ws.cell(row = excel_line,column = 4,value = bytes(read).decode('gb2312'))
            ws.cell(row = excel_line,column = 5,value = result_com)
        elif type == 'configuration_verification' or type == 'default_check':
            if re.match(fv_lines[index][:-1]+'.*',bytes(read).decode('gb2312')):
                result_com = '------PASS'
            else:
                result_com = '------FAIL'
            ws.cell(row = excel_line,column = 1,value = descript)
            ws.cell(row = excel_line,column = 3,value = fv_lines[index][:-1])
            ws.cell(row = excel_line,column = 4,value = bytes(read).decode('gb2312'))
            ws.cell(row = excel_line,column = 5,value = result_com)
        elif type == 'burst_sp':
            if fv_lines[index][:-1] == '0000':
                if(re.match('0000     0.0000     0.0000     0.0000    0.0000    0.0000    0.0000.*',bytes(read).decode('gb2312').split('\n')[1][:-1])):
                    result_com = '------PASS'
                else:
                    result_com = '------FAIL'
            else:
                burst_list = bytes(read).decode('gb2312').split('\n')[1][:-1].split()
                burst_list_int = []
                for cell in burst_list:
                    burst_list_int.append(abs(float(cell)))
                if(0<=burst_list_int[1]<1 and 0<=burst_list_int[2]<1 and 0<=burst_list_int[3]<1 and 0<=burst_list_int[4]<0.1 and 0<=burst_list_int[5]<0.1 and 0.9<burst_list_int[6]<=1.1):
                    result_com = '------PASS'
                else:
                    result_com = '------FAIL'
            ws.cell(row = excel_line,column = 1,value = descript)
            ws.cell(row = excel_line,column = 5,value = result_com)
        elif type == 'power_cycle':
            time.sleep(2)
            result_com = '------PASS'
            ws.cell(row = excel_line,column = 1,value = descript)
            ws.cell(row = excel_line,column = 5,value = result_com)
        else:
            ws.cell(row = excel_line,column = 1,value = descript)
            result_com = '------NO NEED MATCH!'
        if re.match('.*FAIL',result_com):
            fail_num = fail_num + 1
        text_handel.insert(END, current_time+'    '+bytes(read).decode('gb2312')+result_com+'\n')
        f_log.write(current_time+'    '+bytes(read).decode('gb2312')+result_com+'\n')

def message_askokcancel(title, info):
    file_warn = Tk()
    file_warn.withdraw()
    file_warn.update()
    #mes = tkinter.messagebox.askokcancel(title, info)
    tm.showinfo(title = title,message = info)
    file_warn.destroy()
    #return mes

def com_bpsset(com,bps,com1,bps1):
    global comport_new, bpsport_new,comport_new1, bpsport_new1
    comport_new = com.get()
    bpsport_new = bps.get()
    comport_new1 = com1.get()
    bpsport_new1 = bps1.get()
    top.destroy()

def  getcom():
    port_list = list(serial.tools.list_ports.comports())
    serial_com = []
    for index in range(len(port_list)):
        serial_com.append(port_list[index][0])
    return serial_com


def thread_recv(text_handel,data_packets):
    global ser
    read = ser.readline()
    if len(read) > 0:
        text_handel.insert(END, '*')
        data_packets.write(bytes(read).decode('gb2312'))

def thread_send(text_handel):
    global excel_line
    global ser,test_steps,fail_num,result_com
    global filenames,vfilenames
    current_time = time.strftime('%Y_%m_%d_%H_%M_%S',time.localtime(time.time()))
    f = open(filenames)
    f_lines = f.readlines()
    f.close()
    fv = open(vfilenames)
    fv_lines = fv.readlines()
    fv.close()
    f_log =  open((log_filepath+'\\spi_function_test_log_%s.txt')%current_time,'a+')
    wb = Workbook()
    ws = wb.active
    ws.title = 'SPI function test result'
    ws['A1'] = 'Tset name'
    ws['B1'] = 'Valu set'
    ws['C1'] = 'Expected outpt'
    ws['D1'] = 'Actual outpt'
    ws['E1'] = 'Result'
    for index in range(len(f_lines)):
        send_split = f_lines[index].split('#')
        if send_split[0] == 'end of command file':
            current_time = time.strftime('%Y_%m_%d_%H:%M:%S',time.localtime(time.time()))
            text_handel.insert(END, current_time+'    '+'All fail %d steps'%fail_num+'\n')
            f_log.write(current_time+'    '+'All fail %d steps'%fail_num+'\n')
            test_steps = 1
            fail_num = 0
        elif send_split[0] == '1M data packets check with ODR 200Hz':
            if os.path.exists(log_filepath+'\\spi_data_packet.txt'):
                os.remove(log_filepath+'\\spi_data_packet.txt')
            data_packets =  open((log_filepath+'\\spi_data_packet.txt'),'a+')
            ser.write(bytes.fromhex('3E00000A'))
            current_time = time.strftime('%Y_%m_%d_%H:%M:%S',time.localtime(time.time()))
            text_handel.insert(END, current_time+'    ' +'Step%d:'%test_steps+send_split[0]+'\n')
            f_log.write(current_time+'    '+'Step%d:'%test_steps+send_split[0]+'\n')
            test_steps += 1
            time.sleep(0.1)
            for read_index in range(1000):
                thread_recv(text_handel,data_packets)
            data_packets.close()
            packet_1M_check(log_filepath+'\\spi_data_packet.txt')
            text_handel.insert(END,'\n')
            current_time = time.strftime('%Y_%m_%d_%H:%M:%S',time.localtime(time.time()))
            text_handel.insert(END, current_time+'    '+result_com+'\n')
            f_log.write(current_time+'    '+result_com+'\n')
            ws.cell(row = excel_line,column = 1,value = send_split[0])
            ws.cell(row = excel_line,column = 5,value = result_com)
        else:
            default_check(text_handel,send_split[2][:-1],send_split[1],f_log,fv_lines,index,send_split[0],ws)
            excel_line += 1
            
    f_log.close()
    current_time = time.strftime('%Y_%m_%d_%H_%M_%S',time.localtime(time.time()))
    wb.save((log_filepath+'\\spi_result_%s.xlsx')%current_time)
def uut_uart():
    global ser1,bpsport_new1,comport_new1
    ser1 = serial.Serial(port = comport_new1, baudrate = int(bpsport_new1), timeout = 0.2)
    if ser1.is_open:
        pass
    else:
        ser1.open()
    
    ser1.close()

def uart_open(text_handel):
    global ser,bpsport_new,comport_new,filenames
    if re.match('COM.*',comport_new) and re.match('COM.*',comport_new1):
        if ser != 0:
            ser.close()
        ser = serial.Serial(port = comport_new, baudrate = int(bpsport_new), timeout = 0.2)
        if ser.is_open:
            pass
        else:
            ser.open()
        if re.match('.*Cfile.txt',filenames) and re.match('.*Vfile.txt',vfilenames):
            send_data = threading.Thread(target = thread_send,args = (text_handel,),name = 'T2')
            send_data.setDaemon(True)
            send_data.start()
        else:
            file_selectwarn = threading.Thread(target = tm.showwarning,args = ('提示','文件选择错误！'),name = 'T3')
            file_selectwarn.start()
    else:
        port_selectwarn = threading.Thread(target = tm.showwarning,args = ('提示','没有选择串口！'),name = 'T4')
        port_selectwarn.start()

def serial_paraset():
    global top,bpsport_new,comport_new
    com_list = getcom()
    for index in range(len(buadrate_list)):
        if bpsport_new == buadrate_list[index]:
            buadrate_index = index
    
    if com_list == []:
        com_list.append('NO PORT')
        com_index = 0
    else:
        if comport_new in com_list:
            for index_1 in range(len(com_list)):
                if comport_new == com_list[index_1]:
                    com_index = index_1
        else:
            com_index = 0
            
    if len(com_list) > 1:
        com_index1 = 1
    else:
        com_list.append('NO PORT')
        com_index1 = 1
        
    top = Tk()
    top.title('串口设置')
    top.geometry('720x500+150+150')
    label_com = Label(top, text = 'STM32串口：',height = 4).place(x = 20,y = 10,width = 80,height = 40)
    label_com = Label(top, text = '串口号',height = 2).place(x = 20,y = 50,width = 40,height = 30)
    label_bps = Label(top, text = '波特率',height = 2).place(x = 20,y = 86,width = 40,height = 30)
    
    #串口设置
    varport = StringVar()
    combo_com = ttk.Combobox(top, textvariable = varport, width = 8, height = 2, justify = CENTER)
    combo_com['value'] = com_list
    combo_com.place(x = 80, y=50, width = 80, height = 30)
    combo_com.current(com_index)
    
    #波特率设置
    varbitrate = StringVar()
    combo_bps = ttk.Combobox(top, textvariable = varbitrate, width = 8 ,height = 2, justify = CENTER)
    combo_bps['value'] = buadrate_list
    combo_bps.place(x = 80, y = 86, width = 80, height = 30)
    combo_bps.current(buadrate_index)
    
    label_com1 = Label(top, text = 'UUT串口：',height = 4).place(x = 20,y = 150,width = 80,height = 40)
    label_com1 = Label(top, text = '串口号',height = 2).place(x = 20,y = 190,width = 40,height = 30)
    label_bps1 = Label(top, text = '波特率',height = 2).place(x = 20,y = 226,width = 40,height = 30)
    
    #串口设置
    varport1 = StringVar()
    combo_com1 = ttk.Combobox(top, textvariable = varport1, width = 8, height = 2, justify = CENTER)
    combo_com1['value'] = com_list
    combo_com1.place(x = 80, y=190, width = 80, height = 30)
    combo_com1.current(com_index1)
    
    #波特率设置
    varbitrate1 = StringVar()
    combo_bps1 = ttk.Combobox(top, textvariable = varbitrate1, width = 8 ,height = 2, justify = CENTER)
    combo_bps1['value'] = buadrate_list
    combo_bps1.place(x = 80, y = 226, width = 80, height = 30)
    combo_bps1.current(buadrate_index)
    
    make_set = Button(top,text = '确定',width = 18, height = 2)
    make_set.bind('<Button-1>', lambda event : com_bpsset(combo_com,combo_bps,combo_com1,combo_bps1))
    make_set.place(x = 520,y=420)
    
    #top.mainloop()
    

def file_select():
    global filenames
    filenames = tkinter.filedialog.askopenfilename()
    
def vfile_select():
    global vfilenames
    vfilenames = tkinter.filedialog.askopenfilename()
    
    
